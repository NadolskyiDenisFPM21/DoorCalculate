import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import io

from html2excel import ExcelParser


def create(data):
    input_file = 'polls/document_gen/to_excel.html'
    parser = ExcelParser(input_file)
    parser.to_excel('polls/document_gen/to_excel.xlsx')
    wb = openpyxl.load_workbook('polls/document_gen/to_excel.xlsx')
    sheet = wb.active
    
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = Border(
                left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000')
            )
    with io.BytesIO() as buffer:
        wb.save(buffer)
        return buffer.getvalue()


# def create(data):
#     # Создаем новую книгу и активный лист
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active

#     # Записываем заголовки
#     headers = list(data.keys())
#     for col_num, header in enumerate(headers, 1):
#         sheet.column_dimensions[sheet.cell(row=1, column=col_num).column_letter].width = len(header) + 5
#         sheet.cell(row=1, column=col_num, value=header)

#     # Записываем данные
#     for row_num, row_data in enumerate(zip(*data.values()), 2):
#         for col_num, cell_value in enumerate(row_data, 1):
#             sheet.cell(row=row_num, column=col_num, value=cell_value)

#     # Добавляем объект таблицы
#     table = Table(displayName="MyTable", ref=sheet.dimensions)
#     style = TableStyleInfo(
#         name="TableStyleMedium9", showFirstColumn=False,
#         showLastColumn=False, showRowStripes=False, showColumnStripes=False)
#     table.tableStyleInfo = style
#     sheet.add_table(table)


#     with io.BytesIO() as buffer:
#         workbook.save(buffer)
#         return buffer.getvalue()
        
        
    